#!/usr/bin/env python3
"""
build_data.py  -  turn the Pepys Master_Sheet.xlsx into network_data.json for the
web visualisation of the diary network growing over time.

Model (identical to processing_and_analysis/04_network_generator/pepys_dl.py):
  * every "Clean Groups" cell holds one or more groups written  gid:{Name, Name, ...}
  * a group is a set of people who were together on that diary day (an "event")
  * two people are connected when they share a group; an edge's weight is the number
    of groups the pair has shared so far.

This is really a TWO-MODE (affiliation) network: people <-> events. We emit both the
people and the events, so the page can show either the one-mode people projection or
the full two-mode graph.

Layout: a DISTANCE-TIED layout. Positions come from Kamada-Kawai / stress
minimisation on each connected component, so the straight-line distance between two
people approximates their shortest-path (hop) distance in the network. Components are
rescaled to a common "one hop = one unit" scale and then packed with generous gaps so
disconnected groups sit clearly apart.

Output (web/network_data.json):
  meta    : date span, counts, and the named "chapters" used for the jump buttons
  nodes   : every person, with first-appearance date, totals, and a (px, py) position
  events  : chronological list of days; each day lists its groups {id, members}
"""
import re, json, datetime, os, math
from collections import defaultdict
from openpyxl import load_workbook
import networkx as nx

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
INFILE = os.path.join(ROOT, "data", "Master_Sheet.xlsx")
OUTFILE = os.path.join(HERE, "network_data.json")

# --- date parsing (mirrors pepys_dl.py, incl. the +300-year Excel shift) ------
_DATE_FORMATS = [
    "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d",
    "%m/%d/%Y", "%m-%d-%Y", "%m.%d.%Y",
    "%m/%d/%y", "%m-%d-%y", "%m.%d.%y",
    "%d-%b-%Y", "%d-%B-%Y", "%d %b %Y", "%d %B %Y", "%d %b, %Y", "%d %B, %Y",
    "%b %d, %Y", "%B %d, %Y", "%b %d %Y", "%B %d %Y",
]


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return datetime.date(val.year, val.month, val.day)
    s = str(val).strip()
    if not s:
        return None
    s = re.sub(r"[T ]\d{1,2}:\d{2}(:\d{2})?(\.\d+)?\s*$", "", s).strip()
    for f in _DATE_FORMATS:
        try:
            d = datetime.datetime.strptime(s, f)
            return datetime.date(d.year, d.month, d.day)
        except ValueError:
            continue
    return None


# --- read every "Clean Groups" sheet -----------------------------------------
wb = load_workbook(INFILE, read_only=True, data_only=True)

# day -> list of (group_id, [people])   (group id scoped by sheet/year, as in pepys_dl.py)
day_groups = defaultdict(list)
first_seen = {}                    # person -> earliest date
appearances = defaultdict(int)     # person -> number of group memberships

for ws in wb:
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if header is None:
        continue
    head = [str(c).strip().lower() if c else "" for c in header]
    if "clean groups" not in head:
        continue
    gcol = head.index("clean groups")
    dcol = next((i for i, h in enumerate(head) if h.startswith("date")), None)
    sheet_year = int(ws.title) if re.fullmatch(r"\d{4}", ws.title.strip()) else None

    for row in it:
        val = row[gcol] if gcol < len(row) else None
        if not val:
            continue
        raw = row[dcol] if (dcol is not None and dcol < len(row)) else None
        dt = parse_date(raw)
        if dt is not None and sheet_year is not None and dt.year - sheet_year == 300:
            dt = datetime.date(dt.year - 300, dt.month, dt.day)
        if dt is None:
            continue   # undated group rows can't be placed on the timeline
        for gid, names in re.findall(r"(\w+):\{([^}]*)\}", str(val)):
            members = []
            for n in (x.strip() for x in names.split(",")):
                if n and n not in members:      # de-dupe within a single group
                    members.append(n)
            if not members:
                continue
            gid = "%s_%s" % (ws.title, gid)     # scope by sheet (year)
            day_groups[dt].append((gid, members))
            for n in members:
                appearances[n] += 1
                if n not in first_seen or dt < first_seen[n]:
                    first_seen[n] = dt

all_dates = sorted(day_groups)
min_date, max_date = all_dates[0], all_dates[-1]

# --- one-mode weighted person-person graph -----------------------------------
G = nx.Graph()
G.add_nodes_from(first_seen)
pair_weight = defaultdict(int)
for dt in all_dates:
    for gid, members in day_groups[dt]:
        for i in range(len(members)):
            for j in range(i + 1, len(members)):
                a, b = members[i], members[j]
                key = (a, b) if a < b else (b, a)
                pair_weight[key] += 1
for (a, b), w in pair_weight.items():
    G.add_edge(a, b, weight=w)

# --- distance-tied layout, per connected component ---------------------------
# Kamada-Kawai minimises sum (|xi-xj| - hops(i,j))^2, so Euclidean distance tracks
# graph (hop) distance. We run it per component (KK needs a connected graph), then
# rescale every component to the same "one hop ~= one unit" scale.
def component_layout(nodes):
    sub = G.subgraph(nodes)
    m = sub.number_of_nodes()
    if m == 1:
        return {next(iter(nodes)): (0.0, 0.0)}
    if m == 2:
        a, b = list(nodes)
        return {a: (-0.5, 0.0), b: (0.5, 0.0)}
    # unweighted hop distances -> "distance in the network" is number of steps
    pos = nx.kamada_kawai_layout(sub, weight=None)
    pos = {n: (float(p[0]), float(p[1])) for n, p in pos.items()}
    # rescale so the median graph edge spans ~1 unit (a consistent hop scale)
    lens = []
    for a, b in sub.edges():
        dx = pos[a][0] - pos[b][0]
        dy = pos[a][1] - pos[b][1]
        lens.append(math.hypot(dx, dy))
    lens.sort()
    scale = lens[len(lens) // 2] if lens else 1.0
    if scale <= 1e-9:
        scale = 1.0
    return {n: (x / scale, y / scale) for n, (x, y) in pos.items()}


def radius_of(coords):
    cx = sum(x for x, _ in coords) / len(coords)
    cy = sum(y for _, y in coords) / len(coords)
    r = max((math.hypot(x - cx, y - cy) for x, y in coords), default=0.0)
    return cx, cy, r


components = sorted(nx.connected_components(G), key=len, reverse=True)

# extra separation between disconnected components (in hop-units)
COMPONENT_GAP = 5.0

positions = {}     # person -> (x, y) in the shared hop-space

# lay out every component locally and note its radius
laid = []
for comp in components:
    local = component_layout(comp)
    lcx, lcy, r = radius_of(list(local.values()))
    laid.append((local, lcx, lcy, max(r, 0.5)))

# giant component sits at the centre
giant, gcx, gcy, gr = laid[0]
for person, (x, y) in giant.items():
    positions[person] = (x - gcx, y - gcy)

# the small components form compact concentric rings hugging the giant, each ring
# just far enough out to keep a clear COMPONENT_GAP between neighbours
smalls = laid[1:]
if smalls:
    r_small = max(r for _, _, _, r in smalls)
    slot = 2 * r_small + COMPONENT_GAP
    ring = gr + COMPONENT_GAP + r_small
    i = 0
    while i < len(smalls):
        cap = max(1, int((2 * math.pi * ring) / slot))
        batch = smalls[i:i + cap]
        for j, (local, lcx, lcy, r) in enumerate(batch):
            ang = 2 * math.pi * j / len(batch)
            ox, oy = ring * math.cos(ang), ring * math.sin(ang)
            for person, (x, y) in local.items():
                positions[person] = (x - lcx + ox, y - lcy + oy)
        i += cap
        ring += slot

# --- uniform normalise into [0,1]^2 (preserve aspect ratio & distances) -------
xs = [p[0] for p in positions.values()]
ys = [p[1] for p in positions.values()]
minx, maxx, miny, maxy = min(xs), max(xs), min(ys), max(ys)
span = max(maxx - minx, maxy - miny) or 1.0
cx0 = (minx + maxx) / 2.0
cy0 = (miny + maxy) / 2.0


def to_unit(x, y):
    # centre, scale by the SAME factor on both axes, recentre into the unit box
    return (round((x - cx0) / span + 0.5, 4), round((y - cy0) / span + 0.5, 4))


PXY = {p: to_unit(x, y) for p, (x, y) in positions.items()}

degree = dict(G.degree())
nodes = []
for person in sorted(first_seen, key=lambda p: (first_seen[p], p)):
    px, py = PXY[person]
    nodes.append({
        "id": person,
        "first": first_seen[person].isoformat(),
        "count": appearances[person],
        "degree": degree.get(person, 0),
        "px": px, "py": py,
    })

# events: chronological days, each with its groups {id, members} for the two-mode view
events = []
for dt in all_dates:
    groups = [{"i": gid, "m": members} for gid, members in day_groups[dt]]
    events.append({"date": dt.isoformat(), "groups": groups})

n_groups = sum(len(e["groups"]) for e in events)

# --- named chapters for the "jump to" buttons --------------------------------
chapters = [
    {"label": "The Diary Begins", "date": "1664-01-01",
     "note": "Pepys' social world at the start of 1664."},
    {"label": "1665 · The Great Plague", "date": "1665-06-01",
     "note": "Plague grips London through the summer and autumn of 1665."},
    {"label": "1666 · The Great Fire", "date": "1666-09-02",
     "note": "The Great Fire of London breaks out, 2–6 September 1666."},
    {"label": "1667 · Aftermath", "date": "1667-01-01",
     "note": "The final year of this network."},
]
for c in chapters:
    d = datetime.date.fromisoformat(c["date"])
    d = max(min_date, min(max_date, d))
    c["date"] = d.isoformat()

data = {
    "meta": {
        "minDate": min_date.isoformat(),
        "maxDate": max_date.isoformat(),
        "people": len(nodes),
        "days": len(events),
        "groups": n_groups,
        "edges": len(pair_weight),
        "components": len(components),
        "chapters": chapters,
    },
    "nodes": nodes,
    "events": events,
}

with open(OUTFILE, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, separators=(",", ":"))

# also emit a JS wrapper so the page works when opened directly (file://), where
# browsers refuse to fetch() a sibling .json file.
with open(os.path.join(HERE, "network_data.js"), "w", encoding="utf-8") as f:
    f.write("window.PEPYS_DATA = ")
    json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    f.write(";")

print("wrote", OUTFILE)
print("people:", len(nodes), " events(groups):", n_groups, " days:", len(events),
      " edges:", len(pair_weight), " components:", len(components))
print("range:", min_date, "->", max_date)
